# guion_editor/widgets/takeo_dialog.py
import os
import sys
import pandas as pd
import logging
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QFormLayout, QSpinBox, QDialogButtonBox,
    QListWidget, QListWidgetItem, QPushButton, QHBoxLayout, QMessageBox,
    QFileDialog, QAbstractItemView, QGroupBox, QComboBox, QLabel, QApplication,
    QProgressDialog, QCheckBox
)
from PyQt6.QtCore import Qt, QThread
from guion_editor.utils.takeo_optimizer_logic import TakeoWorker, TakeoOptimizerLogic
from guion_editor.widgets.export_selection_dialog import ExportSelectionDialog
from guion_editor import constants as C 
from guion_editor.utils.paths import resource_path

# Importar estilos de OpenPyXL para el formato del Excel
from openpyxl.styles import Font, Alignment, Border, Side

# --- IMPORTACIÓN DINÁMICA DEL CONVERSOR ---
# Intentamos importar asumiendo que se ejecuta desde la raíz.
# Si falla, añadimos la raíz al path relativo a este archivo.
try:
    # Ahora que está dentro de la estructura del paquete, podemos importarlo directamente
    from guion_editor.widgets.xlsx_converter.main import process_excel_to_txt
except ImportError:
    # Fallback por si acaso se ejecuta de forma extraña, intentamos import relativo
    try:
        from .xlsx_converter.main import process_excel_to_txt
    except ImportError as e:
        logging.error(f"No se pudo importar el conversor TXT: {e}")
        process_excel_to_txt = None
# ------------------------------------------

def load_takeo_stylesheet() -> str:
    try:
        # Mucho más limpio y seguro:
        css_path = resource_path(os.path.join('guion_editor', 'styles', 'takeo_styles.css'))
        
        with open(css_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        logging.warning(f"No se pudo cargar takeo_styles.css desde {css_path}: {e}")
        return ""

class TakeoDialog(QDialog):
    def __init__(self, table_window, get_icon_func=None, parent=None):
        super().__init__(parent)
        self.table_window = table_window
        self.get_icon = get_icon_func
        self.setWindowTitle("Optimizador de Takes (Takeo)")
        self.setMinimumSize(600, 550)
        self.setObjectName("TakeoDialog")
        
        # Atributos para el manejo del hilo
        self.thread: QThread | None = None
        self.worker: TakeoWorker | None = None
        
        stylesheet = load_takeo_stylesheet()
        if stylesheet:
            self.setStyleSheet(stylesheet)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        config_group = QGroupBox("Reglas de Takeo")
        form_layout = QFormLayout(config_group)
        
        self.dialogue_source_combo = QComboBox()
        self.dialogue_source_combo.addItems(["EUSKERA", "DIÁLOGO"])
        form_layout.addRow("Columna de diálogo a optimizar:", self.dialogue_source_combo)

        self.duration_spin = QSpinBox(); self.duration_spin.setRange(1, 999); self.duration_spin.setValue(30)
        form_layout.addRow("Duración máxima por take (segundos):", self.duration_spin)
        self.max_lines_spin = QSpinBox(); self.max_lines_spin.setRange(1, 100); self.max_lines_spin.setValue(10)
        form_layout.addRow("Máximo de líneas por take:", self.max_lines_spin)
        self.max_consecutive_spin = QSpinBox(); self.max_consecutive_spin.setRange(1, 100); self.max_consecutive_spin.setValue(5)
        form_layout.addRow("Máx. líneas consecutivas (mismo personaje):", self.max_consecutive_spin)
        self.max_chars_spin = QSpinBox(); self.max_chars_spin.setRange(10, 200); self.max_chars_spin.setValue(60)
        form_layout.addRow("Máx. caracteres por línea (diálogo):", self.max_chars_spin)
        self.silence_spin = QSpinBox(); self.silence_spin.setRange(0, 999); self.silence_spin.setValue(10)
        form_layout.addRow("Máx. silencio entre intervenciones (segundos):", self.silence_spin)
        layout.addWidget(config_group)

        char_group = QGroupBox("Personajes a Incluir en la Optimización")
        char_layout = QVBoxLayout(char_group)
        char_buttons_layout = QHBoxLayout()
        select_all_btn = QPushButton("Seleccionar Todos"); select_all_btn.clicked.connect(self.select_all_characters)
        deselect_all_btn = QPushButton("Deseleccionar Todos"); deselect_all_btn.clicked.connect(self.deselect_all_characters)
        char_buttons_layout.addWidget(select_all_btn); char_buttons_layout.addWidget(deselect_all_btn); char_buttons_layout.addStretch()
        char_layout.addLayout(char_buttons_layout)
        self.char_list_widget = QListWidget(); self.char_list_widget.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.populate_character_list()
        char_layout.addWidget(self.char_list_widget)
        layout.addWidget(char_group)

        # --- Checkbox para generar TXT ---
        self.chk_generate_txt = QCheckBox("Convertir automáticamente el 'Detalle de Takes' a TXT")
        self.chk_generate_txt.setChecked(True) 
        self.chk_generate_txt.setToolTip("Si se marca, se generará un archivo .txt con el formato de grabación basándose en el Excel de detalles generado.")
        layout.addWidget(self.chk_generate_txt)
        # ----------------------------------------

        self.button_box = QDialogButtonBox()
        self.run_button = self.button_box.addButton("Optimizar y Guardar...", QDialogButtonBox.ButtonRole.AcceptRole)
        cancel_button = self.button_box.addButton(QDialogButtonBox.StandardButton.Cancel)
        self.run_button.clicked.connect(self.run_optimization)
        cancel_button.clicked.connect(self.reject)
        layout.addWidget(self.button_box)

    def populate_character_list(self):
        self.char_list_widget.clear()
        characters = self.table_window.get_character_names_from_model()
        for char in characters:
            item = QListWidgetItem(char); item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked); self.char_list_widget.addItem(item)

    def select_all_characters(self):
        for i in range(self.char_list_widget.count()): self.char_list_widget.item(i).setCheckState(Qt.CheckState.Checked)

    def deselect_all_characters(self):
        for i in range(self.char_list_widget.count()): self.char_list_widget.item(i).setCheckState(Qt.CheckState.Unchecked)

    def get_selected_characters(self):
        return [self.char_list_widget.item(i).text() for i in range(self.char_list_widget.count()) if self.char_list_widget.item(i).checkState() == Qt.CheckState.Checked]

    def run_optimization(self):
        if self.thread and self.thread.isRunning():
            logging.warning("Intento de iniciar optimización mientras una ya está en curso.")
            return

        df = self.table_window.pandas_model.dataframe()
        if not df.empty and C.COL_PERSONAJE in df.columns:
            dirty_mask = df[C.COL_PERSONAJE].astype(str) != df[C.COL_PERSONAJE].astype(str).str.strip()
            if dirty_mask.any():
                dirty_rows = df[dirty_mask]
                unique_dirty_names = sorted(dirty_rows[C.COL_PERSONAJE].unique())
                count = len(unique_dirty_names)
                msg = (f"Se han encontrado {count} personajes con espacios sobrantes.\nSeleccione una acción:")
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("Error de Formato en Personajes")
                msg_box.setText(msg)
                msg_box.setIcon(QMessageBox.Icon.Warning)
                btn_fix = msg_box.addButton("Corregir Automáticamente", QMessageBox.ButtonRole.AcceptRole)
                btn_export = msg_box.addButton("Exportar Excel (Debug)", QMessageBox.ButtonRole.ActionRole)
                btn_cancel = msg_box.addButton(QMessageBox.StandardButton.Cancel)
                msg_box.exec()
                clicked_button = msg_box.clickedButton()
                if clicked_button == btn_fix: self._perform_cleanup()
                elif clicked_button == btn_export: self._export_dirty_rows(dirty_rows); return
                else: return

        config = {'max_duration': self.duration_spin.value(), 'max_lines_per_take': self.max_lines_spin.value(), 'max_consecutive_lines_per_character': self.max_consecutive_spin.value(), 'max_chars_per_line': self.max_chars_spin.value(), 'max_silence_between_interventions': self.silence_spin.value()}
        selected_chars = self.get_selected_characters()
        if not selected_chars: 
            QMessageBox.warning(self, "Sin Selección", "Debe seleccionar al menos un personaje.")
            return
        
        current_df = self.table_window.pandas_model.dataframe()
        dialogue_col = self.dialogue_source_combo.currentText()
        
        # Extraer mapa de reparto si existe
        reparto_map = {}
        if C.COL_REPARTO in current_df.columns:
            temp_df = current_df[[C.COL_PERSONAJE, C.COL_REPARTO]].dropna()
            for _, row in temp_df.iterrows():
                char = str(row[C.COL_PERSONAJE]).strip()
                actor = str(row[C.COL_REPARTO]).strip()
                if char and actor: reparto_map[char] = actor
        
        self.thread = QThread()
        self.worker = TakeoWorker(config, current_df.copy(), selected_chars, dialogue_col, reparto_map)
        self.worker.moveToThread(self.thread)
        self.worker.finished.connect(self._on_optimization_finished)
        self.worker.error.connect(self._on_optimization_error)
        self.thread.started.connect(self.worker.run)
        self.thread.finished.connect(self.thread.deleteLater)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        
        self.thread.start()
        self.run_button.setEnabled(False)
        self.run_button.setText("Procesando...")
        self.setCursor(Qt.CursorShape.WaitCursor)

    def _export_dirty_rows(self, dirty_rows_df: pd.DataFrame):
        if dirty_rows_df.empty: return
        filename = "DEBUG_Personajes_Con_Espacios.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Exportar Reporte de Errores", filename, "Archivos Excel (*.xlsx)")
        if path:
            try: dirty_rows_df.to_excel(path, index=False); QMessageBox.information(self, "Exportación Exitosa", f"Guardado en:\n{path}")
            except Exception as e: QMessageBox.critical(self, "Error", f"Error: {e}")

    def _perform_cleanup(self):
        selected_stripped = []
        for i in range(self.char_list_widget.count()):
            item = self.char_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked: selected_stripped.append(item.text().strip())
        self.table_window.trim_all_character_names()
        self.populate_character_list()
        for i in range(self.char_list_widget.count()):
            item = self.char_list_widget.item(i)
            if item.text() in selected_stripped: item.setCheckState(Qt.CheckState.Checked)
            else: item.setCheckState(Qt.CheckState.Unchecked)

    def _on_optimization_finished(self, detail_df, summary_df, failures_df, problematic_report, takes_generated):
        logging.info("Recepción de resultados del worker de Takeo.")
        self.unsetCursor()
        self.run_button.setEnabled(True)
        self.run_button.setText("Optimizar y Guardar...")
        self.save_results(detail_df, summary_df, failures_df, problematic_report, takes_generated)
        self.accept()

    def _on_optimization_error(self, error_string):
        self.unsetCursor()
        self.run_button.setEnabled(True)
        self.run_button.setText("Optimizar y Guardar...")
        QMessageBox.critical(self, "Error en la Optimización", f"Ocurrió un error durante el proceso:\n{error_string}")
        self.reject()

    def save_formatted_summary(self, path: str, df: pd.DataFrame, title_text: str):
        """Guarda el resumen con formato visual usando OpenPyXL."""
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=1, sheet_name='Resumen')
            workbook = writer.book
            worksheet = writer.sheets['Resumen']
            
            # Título
            worksheet['A1'] = title_text
            max_col = len(df.columns)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title_cell = worksheet['A1']
            title_cell.font = Font(bold=True, size=14, name='Calibri')
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Estilos de tabla
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 2: # Cabecera
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column == 2: # Columna B (Takes) alineada a derecha
                        cell.alignment = Alignment(horizontal='right')
            
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 30

    def save_results(self, detail_df, summary_df, failures_df, problematic_report, takes_generated):
            reports_available = {
                "detail": not detail_df.empty, 
                "summary": not summary_df.empty, 
                "failures": not failures_df.empty, 
                "problems": bool(problematic_report)
            }
            if not any(reports_available.values()):
                QMessageBox.information(self, "Proceso Completado", "La optimización finalizó, pero no se generaron resultados para exportar."); return

            export_dialog = ExportSelectionDialog(reports_available, self);
            if not export_dialog.exec(): return

            choices = export_dialog.get_choices()
            if not choices or not any(choices.values()):
                QMessageBox.information(self, "Sin Selección", "No se ha seleccionado ningún archivo para exportar."); return

            save_dir = QFileDialog.getExistingDirectory(self, "Seleccionar Directorio para Guardar Informes")
            if not save_dir: return

            # --- RECOPILAR DATOS DE CABECERA DE LA UI PRINCIPAL ---
            header_data = self.table_window._get_header_data_from_ui()
            product = str(header_data.get("product_name", "")).strip()
            chapter = str(header_data.get("chapter_number", "")).strip()
            
            # Título para el Excel de Resumen
            title_text_excel = f"{product.upper()} {chapter}".strip()
            if not title_text_excel: title_text_excel = "RESUMEN DE TAKES"

            saved_files = []
            detail_excel_path = None

            try:
                if choices["detail"] and reports_available["detail"]:
                    path = os.path.join(save_dir, "detalle_takes_optimizado.xlsx")
                    detail_df.to_excel(path, index=False)
                    saved_files.append(os.path.basename(path))
                    detail_excel_path = path # Guardamos la ruta para la conversión
                
                if choices["summary"] and reports_available["summary"]:
                    path = os.path.join(save_dir, "resumen_takes_optimizado.xlsx")
                    self.save_formatted_summary(path, summary_df, title_text_excel)
                    saved_files.append(os.path.basename(path))
                    
                if choices["failures"] and reports_available["failures"]:
                    path = os.path.join(save_dir, "reporte_fallos_de_agrupacion.xlsx"); failures_df.to_excel(path, index=False); saved_files.append(os.path.basename(path))
                if choices["problems"] and reports_available["problems"]:
                    prob_df = pd.DataFrame(problematic_report); path = os.path.join(save_dir, "reporte_intervenciones_problematicas.xlsx"); prob_df.to_excel(path, index=False); saved_files.append(os.path.basename(path))

                # --- CONVERSIÓN AUTOMÁTICA A TXT MEJORADA ---
                if self.chk_generate_txt.isChecked() and detail_excel_path and process_excel_to_txt:
                    try:
                        target_col = self.dialogue_source_combo.currentText()
                        
                        # 1. Construir nombre de archivo: NombreProducto_DIALOG.txt
                        # Si no hay nombre de producto, usar un genérico "Guion_DIALOG.txt"
                        clean_product_name = product.replace(" ", "_")
                        if not clean_product_name:
                            clean_product_name = "Guion"
                        
                        txt_filename = f"{clean_product_name}_DIALOG.txt"

                        # 2. Llamar al conversor pasando datos extra
                        txt_path = process_excel_to_txt(
                            detail_excel_path, 
                            target_col,
                            header_data=header_data,        # Pasamos Título y Capítulo
                            custom_output_name=txt_filename # Pasamos el nombre forzado
                        )
                        saved_files.append(os.path.basename(txt_path))
                        
                    except Exception as ex_conv:
                        logging.error(f"Error al generar TXT automático: {ex_conv}")
                        QMessageBox.warning(self, "Advertencia TXT", f"Los Excels se generaron, pero falló la conversión automática a TXT:\n{ex_conv}")
                elif self.chk_generate_txt.isChecked() and not process_excel_to_txt:
                    QMessageBox.warning(self, "Advertencia", "El módulo de conversión TXT no se pudo cargar, por lo que no se generó el archivo de texto.")
                # -----------------------------------

                if not saved_files:
                    QMessageBox.information(self, "Sin Resultados", "No se guardó ningún archivo según su selección."); return
                
                sum_val = summary_df.iloc[-1]["TAKES (apariciones)"] if reports_available["summary"] else "N/A"
                files_str = "\n - ".join(saved_files)
                QMessageBox.information(self, "Proceso Completado", f"Optimización finalizada.\n\nTakes únicos generados: {takes_generated}\nSuma de apariciones: {sum_val}\n\nArchivos guardados en '{os.path.basename(save_dir)}':\n - {files_str}")
            except Exception as e:
                logging.error("No se pudieron guardar los informes de Takeo.", exc_info=True)
                QMessageBox.critical(self, "Error al Guardar", f"No se pudieron guardar los informes: {e}")
            
    def closeEvent(self, event):
        if self.thread and self.thread.isRunning():
            logging.info("Cerrando diálogo de Takeo, intentando detener el hilo...")
            self.thread.quit()
            self.thread.wait(500) 
        super().closeEvent(event)