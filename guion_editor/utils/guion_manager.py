# guion_editor/utils/guion_manager.py
import pandas as pd
import json
import os
from typing import Tuple, Dict, Any
from openpyxl.styles import PatternFill

from .dialog_utils import leer_guion
from guion_editor import constants_logic as C

class GuionManager:
    BASE_COLUMNS = [C.COL_IN, C.COL_OUT, C.COL_PERSONAJE, C.COL_DIALOGO]
    ALL_COLUMNS = C.DF_COLUMN_ORDER

    def __init__(self) -> None:
        pass

    def process_dataframe(self, df: pd.DataFrame, file_source: str = "unknown") -> Tuple[pd.DataFrame, bool]:
        missing_cols = [col for col in self.BASE_COLUMNS if col not in df.columns]
        if missing_cols:
            pass

        if C.COL_ID not in df.columns:
            if not df.empty:
                df.insert(0, C.COL_ID, range(len(df)))
            else:
                df[C.COL_ID] = pd.Series(dtype='int')

        has_scene_numbers = False
        if C.COL_SCENE not in df.columns:
            insert_idx_for_scene = df.columns.get_loc(C.COL_ID) + 1 if C.COL_ID in df.columns else 0
            df.insert(insert_idx_for_scene, C.COL_SCENE, "1")
        else:
            df[C.COL_SCENE] = df[C.COL_SCENE].astype(str).str.strip().replace(['', 'nan', 'none', 'NaN', 'None'], pd.NA).ffill().fillna("1")

            def normalize_scene_value(scene_str: str) -> str:
                scene_str_stripped = scene_str.strip()
                if scene_str_stripped.endswith(".0"):
                    potential_int_part = scene_str_stripped[:-2]
                    try:
                        if int(potential_int_part) == float(scene_str_stripped):
                            return potential_int_part
                    except ValueError:
                        pass
                return scene_str_stripped
            
            df[C.COL_SCENE] = df[C.COL_SCENE].apply(normalize_scene_value)
            non_empty_scenes = [s for s in df[C.COL_SCENE].tolist() if s.strip() and s.strip().lower() != 'nan']
            has_scene_numbers = not non_empty_scenes or len(set(non_empty_scenes)) > 1 or (len(set(non_empty_scenes)) == 1 and list(set(non_empty_scenes))[0] != "1")

        if C.COL_EUSKERA not in df.columns:
            insert_pos = df.columns.get_loc(C.COL_DIALOGO) + 1 if C.COL_DIALOGO in df.columns else len(df.columns)
            df.insert(insert_pos, C.COL_EUSKERA, "")

        if C.COL_OHARRAK not in df.columns:
            insert_pos_oh = df.columns.get_loc(C.COL_EUSKERA) + 1 if C.COL_EUSKERA in df.columns else len(df.columns)
            df.insert(insert_pos_oh, C.COL_OHARRAK, "")
        
        if C.COL_BOOKMARK not in df.columns:
            df[C.COL_BOOKMARK] = False
        else:
            df[C.COL_BOOKMARK] = df[C.COL_BOOKMARK].fillna(False).astype(bool)

        ordered_present_columns = [col for col in self.ALL_COLUMNS if col in df.columns]
        extra_cols = [col for col in df.columns if col not in self.ALL_COLUMNS]
        df = df[ordered_present_columns + extra_cols]
        return df, has_scene_numbers

    def check_excel_columns(self, path: str) -> Tuple[pd.DataFrame, Dict[str, Any], bool]:
        try:
            xls = pd.ExcelFile(path)
            df = pd.read_excel(xls, sheet_name=0)
            header_data = {}
            if 'Header' in xls.sheet_names:
                try:
                    header_df = pd.read_excel(xls, sheet_name='Header', header=None, index_col=0)
                    if not header_df.empty:
                         header_data = header_df[1].to_dict()
                         for key in ["reference_number", "chapter_number"]:
                             if key in header_data and pd.notna(header_data[key]):
                                 header_data[key] = str(int(header_data[key])) if isinstance(header_data[key], (int, float)) else str(header_data[key])
                except Exception:
                    pass
            
            expected_cols_in_excel = [col for col in self.ALL_COLUMNS if col not in [C.COL_ID, C.COL_BOOKMARK]]
            needs_mapping = not all(col in df.columns for col in expected_cols_in_excel)
            return df, header_data, needs_mapping
        except Exception as e:
            raise

    def save_to_excel(self, path: str, dataframe: pd.DataFrame, header_data: Dict[str, Any]) -> None:
        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df_to_save = dataframe.copy()
                def replace_if_empty(value):
                    return "" if pd.isna(value) or str(value).strip() == '' else value
                
                for col in [C.COL_DIALOGO, C.COL_EUSKERA, C.COL_OHARRAK, C.COL_REPARTO]:
                    if col in df_to_save.columns:
                        df_to_save[col] = df_to_save[col].apply(replace_if_empty)
                        
                df_to_save.to_excel(writer, sheet_name='Guion', index=False)
                if C.COL_OHARRAK in df_to_save.columns:
                    workbook = writer.book
                    worksheet = writer.sheets['Guion']
                    highlight_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                    for df_index, row in df_to_save[df_to_save[C.COL_OHARRAK].astype(str).str.strip() != ''].iterrows():
                        excel_row_index = df_index + 2
                        for col_idx in range(1, len(df_to_save.columns) + 1):
                            worksheet.cell(row=excel_row_index, column=col_idx).fill = highlight_fill
                if header_data:
                    header_df = pd.DataFrame(list(header_data.items()), columns=['Campo', 'Valor'])
                    header_df.to_excel(writer, sheet_name='Header', index=False)
        except Exception as e:
            raise

    def load_from_json(self, path: str) -> Tuple[pd.DataFrame, Dict[str, Any], bool]:
        try:
            with open(path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            header_data = json_data.get("header", {})
            script_data = json_data.get("data", [])
            df = pd.DataFrame(script_data)
            df_processed, has_scenes = self.process_dataframe(df, file_source=path)
            return df_processed, header_data, has_scenes
        except Exception as e:
            raise

    def save_to_json(self, path: str, dataframe: pd.DataFrame, header_data: Dict[str, Any]) -> None:
        try:
            df_to_save = dataframe.copy()
            data_records = df_to_save.to_dict(orient='records')
            json_output = {"header": header_data, "data": data_records}
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, ensure_ascii=False, indent=4)
        except Exception as e:
            raise

    def _convert_tc_to_srt_format(self, tc: str) -> str:
        try:
            parts = tc.split(':')
            h, m, s, f = map(int, parts)
            ms_total = (h * 3600 + m * 60 + s) * 1000 + int(round((f / C.FPS) * 1000.0))
            s_total, mmm = divmod(ms_total, 1000)
            h_srt, s_rem = divmod(s_total, 3600)
            m_srt, s_srt = divmod(s_rem, 60)
            return f"{int(h_srt):02d}:{int(m_srt):02d}:{int(s_srt):02d},{int(mmm):03d}"
        except Exception:
            return "00:00:00,000"

    def save_to_srt(self, path: str, dataframe: pd.DataFrame, column_to_export: str = C.COL_DIALOGO) -> None:
        if column_to_export not in dataframe.columns:
            raise ValueError(f"La columna '{column_to_export}' no se encuentra en el guion.")
        try:
            with open(path, 'w', encoding='utf-8') as f:
                srt_index = 1
                for _, row in dataframe.iterrows():
                    in_tc, out_tc = row[C.COL_IN], row[C.COL_OUT]
                    dialogue = str(row[column_to_export]).strip()
                    if dialogue and pd.notna(in_tc) and pd.notna(out_tc):
                        start_srt = self._convert_tc_to_srt_format(in_tc)
                        end_srt = self._convert_tc_to_srt_format(out_tc)
                        f.write(f"{srt_index}\n{start_srt} --> {end_srt}\n{dialogue}\n\n")
                        srt_index += 1
        except Exception as e:
            raise

    def load_from_docx(self, path: str) -> Tuple[pd.DataFrame, Dict[str, Any], bool]:
        try:
            guion_list_of_dicts = leer_guion(path)
            df = pd.DataFrame(guion_list_of_dicts) if guion_list_of_dicts else pd.DataFrame(columns=self.BASE_COLUMNS)
            df_processed, _ = self.process_dataframe(df, file_source=path)
            return df_processed, {}, False 
        except Exception as e:
            raise